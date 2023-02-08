from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from livros.models import Livro, Categoria
from openpyxl.styles import PatternFill
import re 

class Command(BaseCommand):
    help = 'Analisa um excel e cria livros e categorias no banco de dados'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='O caminho do arquivo excel')

        # Quando você usa o comando python manage.py populate_db -add, ele vai analisar os livros 
        # independentemente deles possuírem o campo Adicionado como Sim ou Não
        parser.add_argument('-add', action='store_true', help='Levar em consideração os livros já adicionados')

    def handle(self, *args, **options):

        # Carrega o arquivo excel
        self.stdout.write("Loading {}...".format(options['file_path']))
        wb = load_workbook(options['file_path'])

        for sheet in wb.worksheets:

            # cria um padrão de cor vermelha para os livros que não são adicionados
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            # cria um padrão de cor preta para os livros que não tem código
            black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

            # Criar um padrão de cor vermelha escuro para os dados obrigatórios que não foram preenchidos
            dark_red_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")

            # Cira um padrão de cor verde para os livros que foram adicionados
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

            # Cria uma categoria com o nome da aba e a descrição vazia, checando se ela já existe
            categoria_name = sheet.cell(row=1, column=1).value
            categoria, categoria_created = Categoria.objects.get_or_create(nome=categoria_name, defaults={'descricao': ""})

            if categoria_created:
                self.stdout.write(self.style.SUCCESS("Categoria {} criada".format(categoria_name)))
            else:   
                self.stdout.write(self.style.SUCCESS("Categoria {} já existe".format(categoria_name)))

            # Itera sobre as linhas do excel começando da linha 3
            for i in range(3, sheet.max_row + 1):

                row = sheet[i]

                # se a linha estiver vazia ela é ignorada
                if not any([cell.value for cell in row]):
                    break


                # Valores que não são aceitos no banco de dados
                INVALID_VALUES = ["", "-", "Incompleto", "Apostila", " "]

                # Dicionário que vai conter os dados do livro
                book_dict = {}	
                book_dict['Categoria'] = categoria
                book_dict['Disponivel'], book_dict['Titulo'], book_dict['Codigo'], book_dict['Autores'], book_dict['Volume'], book_dict['Edicao'], book_dict['Estado'], book_dict['Tipo'], book_dict['Adicionado'] = [cell.value for cell in row[:9]]


                # Se o usuário não usar o argumento -add, os livros que já estiverem no banco de dados serão ignorados
                if not options['add']:
                    if book_dict['Adicionado'] == "Sim":
                        self.stdout.write(self.style.SUCCESS("Livro {} já existe (not -add)".format(book_dict['Codigo'])))
                        continue

                
                # Se algum dado for inválido, ele é substituído por None
                for key, value in book_dict.items():
                    if value in INVALID_VALUES:
                        book_dict[key] = None
               

               # Se o livro não tiver código, ele é ignorado e a célula é pintada de preto e o restante das células da linha é pintada de vermelho
                if book_dict['Codigo'] is None:
                    self.stdout.write(self.style.ERROR("Livro {} não tem código".format(book_dict['Titulo'])))
                    for cell in row[:9]:
                        cell.fill = red_fill
                    row[2].fill = black_fill
                    row[8].value = "Não"
                    continue
                
                # cria um livro com os dados do dicionário, checando se ele já existe
                book, created = Livro.objects.get_or_create(Codigo=book_dict['Codigo'].upper(), defaults={
                    'Disponivel': book_dict['Disponivel'] == "Sim" if book_dict['Disponivel'] else None,
                    'Titulo': book_dict['Titulo'].title() if book_dict['Titulo'] else None,
                    'Autores': book_dict['Autores'].title() if book_dict['Autores'] else None,
                    'Volume': 1 if book_dict['Volume'] == "Único" else int(book_dict['Volume']) if book_dict['Volume'] else None,
                    'Edicao': int(re.sub(r"[^\d]", "", book_dict['Edicao'])) if book_dict['Edicao'] else None,
                    'Estado': book_dict['Estado'].capitalize() if book_dict['Estado'] else None,
                    'Tipo': book_dict['Tipo'].capitalize() if book_dict['Tipo'] else None,
                    'Categoria': book_dict['Categoria'] if book_dict['Categoria'] else None,
                })


                # Se o livro for criado, checa se os dados obrigatórios foram preenchidos, se não foram, a célula é pintada de vermelho escuro
                # e o restante das células da linha é pintada de vermelho
                if created:    
                    row[8].value = "Sim"

                    self.stdout.write(self.style.SUCCESS("Livro {} criado".format(book_dict['Codigo']))) 
                    self.stdout.write("Titulo: {}".format(book.Titulo))	
                    self.stdout.write("Codigo: {}".format(book.Codigo))
                    self.stdout.write("Autores: {}".format(book.Autores))
                    self.stdout.write("Volume: {}".format(book.Volume))
                    self.stdout.write("Edicao: {}".format(book.Edicao))
                    self.stdout.write("Estado: {}".format(book.Estado))
                    self.stdout.write("Tipo: {}".format(book.Tipo))
                    self.stdout.write("Categoria: {}".format(book.Categoria))
                    self.stdout.write("Disponivel: {}".format(book.Disponivel))
                    

                    if not all([book.Categoria, book.Disponivel != None , book.Titulo, book.Codigo, book.Autores, book.Estado, book.Tipo]):
                        Livro.objects.filter(Codigo=book_dict['Codigo'].upper()).delete()
                        self.stdout.write(self.style.ERROR("Livro {} foi deletado".format(book_dict['Codigo'])))
                        row[8].value = "Não"
                        
                        for cell in row[:9]:
                            cell.fill = red_fill
                        
                        
                        if not book.Disponivel != None:
                            self.stdout.write(self.style.ERROR("book.Disponivel: {}".format(book.Disponivel)))
                            row[0].fill = dark_red_fill
                        
                        if not book.Titulo:
                            self.stdout.write(self.style.ERROR("book.Titulo: {}".format(book.Titulo)))
                            row[1].fill = dark_red_fill
                        
                        
                        if not book.Autores:
                            self.stdout.write(self.style.ERROR("book.Autores: {}".format(book.Autores)))
                            row[3].fill = dark_red_fill

                        if not book.Estado:
                            self.stdout.write(self.style.ERROR("book.Estado: {}".format(book.Estado)))
                            row[6].fill = dark_red_fill
                        
                        if not book.Tipo:
                            self.stdout.write(self.style.ERROR("book.Tipo: {}".format(book.Tipo)))
                            row[7].fill = dark_red_fill
                else:
                    self.stdout.write(self.style.SUCCESS("Livro {} já existe".format(book_dict['Codigo'])))
                    row[8].value = "Sim"

                if row[8].value == "Sim":
                    for cell in row[:9]:
                        cell.fill = green_fill
                    

        # Salva o arquivo atualizado no mesmo diretório do arquivo original com o nome original + "_atualizado.xlsx"
        wb.save(options['file_path'][0:-5] + "_atualizado.xlsx")

